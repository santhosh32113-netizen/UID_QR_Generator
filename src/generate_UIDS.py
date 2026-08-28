#!/usr/bin/env python3
"""
Generate deterministic KUIN-G values and QR codes from an Excel workbook.

Input:
    Column B header must be "Drone ID".
    Data starts at row 3.

Outputs:
    Master register:
        Drone ID | KUIN-G

    Distributable register:
        KUIN-G | QR Code
        The QR image is embedded in the workbook.

    QR directory:
        One PNG file per KUIN-G, named <KUIN-G>.png.
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import unicodedata
from pathlib import Path

import qrcode
from qrcode.image.svg import SvgPathImage
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment

INPUT_HEADER = "Drone ID"
KUIN_HEADER = "KUIN-G"
QR_HEADER = "QR Code"


def normalize_existing_id(value: object) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKC", text)
    text = " ".join(text.split())
    return text.casefold()


def make_kuin(drone_id: object) -> str:
    digest = hashlib.sha256(str(drone_id).encode("utf-8")).digest()
    return base64.b32encode(digest).decode("ascii").rstrip("=")[:16]


def validate_input_workbook(input_path: Path) -> list[object]:
    wb = load_workbook(input_path, read_only=True, data_only=True)

    try:
        ws = wb.active

        if ws["B2"].value != INPUT_HEADER:
            raise ValueError(
                f'Input workbook B2 must contain "{INPUT_HEADER}". '
                f"Found: {ws['B2'].value!r}"
            )

        values = []

        for row in range(3, ws.max_row + 1):
            value = ws.cell(row=row, column=2).value

            if value is None or str(value).strip() == "":
                continue

            values.append(value)

        if not values:
            raise ValueError("No Drone ID values were found below B2.")

        return values
    finally:
        wb.close()


def ensure_unique_ids(values: list[object]) -> None:
    seen: dict[str, int] = {}

    for row_number, value in enumerate(values, start=2):
        normalized = normalize_existing_id(value)

        if normalized in seen:
            raise ValueError(
                f"Duplicate Drone ID after normalization: {value!r}. "
                f"First occurrence was row {seen[normalized]}; "
                f"duplicate is row {row_number}."
            )

        seen[normalized] = row_number


def style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    ws.row_dimensions[1].height = 24


def write_master_workbook(output_path: Path, values, kuin_values) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Register"

    ws["A1"] = INPUT_HEADER
    ws["B1"] = KUIN_HEADER

    for row_number, (existing_id, kuin) in enumerate(zip(values, kuin_values), start=2):
        ws.cell(row=row_number, column=1, value=existing_id)
        ws.cell(row=row_number, column=2, value=kuin)

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 28

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def create_qr_png(kuin: str, output_path: Path) -> None:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(kuin)
    qr.make(fit=True)

    image = qr.make_image()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_path, format="PNG")


def create_qr_svg(kuin: str, output_path: Path) -> None:
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
        image_factory=SvgPathImage,
    )
    qr.add_data(kuin)
    qr.make(fit=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    qr.make_image().save(output_path)


def qr_matrix(kuin: str):
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M, border=4)
    qr.add_data(kuin)
    qr.make(fit=True)
    return qr.get_matrix()


def create_qr_stl(kuin: str, output_path: Path) -> None:
    matrix = qr_matrix(kuin)
    module_size = 1.0
    base_height = 0.4
    module_height = 0.6
    size = len(matrix) * module_size
    triangles = []

    def add_box(x, y, width, height):
        vertices = [
            (x, y, base_height), (x + width, y, base_height),
            (x + width, y + width, base_height), (x, y + width, base_height),
            (x, y, height), (x + width, y, height),
            (x + width, y + width, height), (x, y + width, height),
        ]
        faces = [(0, 1, 2), (0, 2, 3), (4, 6, 5), (4, 7, 6),
                 (0, 4, 5), (0, 5, 1), (1, 5, 6), (1, 6, 2),
                 (2, 6, 7), (2, 7, 3), (3, 7, 4), (3, 4, 0)]
        triangles.extend([(vertices[a], vertices[b], vertices[c]) for a, b, c in faces])

    add_box(0, 0, size, base_height)
    for row, values in enumerate(matrix):
        for column, dark in enumerate(values):
            if dark:
                add_box(column * module_size, (len(matrix) - row - 1) * module_size, module_size, base_height + module_height)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="ascii") as stl:
        stl.write(f"solid KUIN_G_{kuin}\n")
        for first, second, third in triangles:
            stl.write("  facet normal 0 0 0\n    outer loop\n")
            for vertex in (first, second, third):
                stl.write(f"      vertex {vertex[0]:.4f} {vertex[1]:.4f} {vertex[2]:.4f}\n")
            stl.write("    endloop\n  endfacet\n")
        stl.write(f"endsolid KUIN_G_{kuin}\n")


def write_qr_register(output_path: Path, records) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["Drone ID", "KUIN-G", "QR PNG File", "QR SVG File", "QR STL File"])
        writer.writeheader()
        writer.writerows(records)


def remove_orphan_qr_files(qr_dir: Path, kuin_values) -> None:
    expected = {f"{kuin}.{extension}" for kuin in kuin_values for extension in ("png", "svg")}
    for path in qr_dir.iterdir():
        if path.suffix.lower() not in {".png", ".svg"}:
            continue
        if path.name not in expected:
            path.unlink()


def remove_orphan_stl_files(stl_dir: Path, kuin_values) -> None:
    expected = {f"{kuin}.stl" for kuin in kuin_values}
    for path in stl_dir.glob("*.stl"):
        if path.name not in expected:
            path.unlink()


def write_qr_stl_backup(stl_dir: Path, kuin_values) -> None:
    stl_dir.mkdir(parents=True, exist_ok=True)
    for kuin in kuin_values:
        create_qr_stl(kuin, stl_dir / f"{kuin}.stl")
    remove_orphan_stl_files(stl_dir, kuin_values)


def write_distributable_workbook(output_path: Path, kuin_values, qr_dir: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distributable Register"

    ws["A1"] = KUIN_HEADER
    ws["B1"] = QR_HEADER

    for row_number, kuin in enumerate(kuin_values, start=2):
        ws.cell(row=row_number, column=1, value=kuin)

        qr_path = qr_dir / f"{kuin}.png"
        create_qr_png(kuin, qr_path)
        create_qr_svg(kuin, qr_dir / f"{kuin}.svg")

        img = XLImage(str(qr_path))
        img.width = 150
        img.height = 150
        img.anchor = f"B{row_number}"
        ws.add_image(img)

        ws.row_dimensions[row_number].height = 115

    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{len(kuin_values) + 1}"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def generate(
    input_path: Path,
    master_output: Path,
    distributable_output: Path,
    qr_dir: Path,
) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    values = validate_input_workbook(input_path)

    ensure_unique_ids(values)

    kuin_values = [make_kuin(value) for value in values]

    write_master_workbook(master_output, values, kuin_values)
    write_distributable_workbook(distributable_output, kuin_values, qr_dir)
    remove_orphan_qr_files(qr_dir, kuin_values)
    stl_dir = qr_dir.parent / "qr_stl_backup"
    write_qr_stl_backup(stl_dir, kuin_values)
    write_qr_register(
        distributable_output.parent / "qr_register.csv",
                [{"Drone ID": drone_id, "KUIN-G": kuin,
                      "QR PNG File": f"{kuin}.png", "QR SVG File": f"{kuin}.svg",
                      "QR STL File": f"{kuin}.stl"}
         for drone_id, kuin in zip(values, kuin_values)],
    )

    print()
    print("KUIN-G generation completed successfully.")
    print(f"Records processed:       {len(values)}")
    print(f"Master register:         {master_output.resolve()}")
    print(f"Distributable register:  {distributable_output.resolve()}")
    print(f"QR directory:            {qr_dir.resolve()}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate deterministic KUIN-G values and QR codes from Excel."
    )
    parser.add_argument("--input", required=True, help="Input Excel workbook.")
    parser.add_argument("--master-output", required=True, help="Master Excel output.")
    parser.add_argument(
        "--distributable-output",
        required=True,
        help="Distributable Excel output.",
    )
    parser.add_argument(
        "--qr-dir",
        required=True,
        help="Directory where QR PNG files will be written.",
    )

    args = parser.parse_args()

    try:
        generate(
            input_path=Path(args.input),
            master_output=Path(args.master_output),
            distributable_output=Path(args.distributable_output),
            qr_dir=Path(args.qr_dir),
        )
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

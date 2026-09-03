#!/usr/bin/env python3
"""Quick health check for the KUIN-G local database and QR pipeline."""
from pathlib import Path
import os
import sys
from openpyxl import load_workbook

ROOT = Path(os.environ.get("KUIN_G_DATA_ROOT", Path(__file__).resolve().parents[1]))
WORKBOOK = ROOT / "input" / "Sample.xlsx"
QR_DIR = ROOT / "qr_codes"

print(f"Data root : {ROOT}")
print(f"Workbook  : {WORKBOOK}")
print(f"Writable  : {os.access(ROOT, os.W_OK)}")
print(f"Workbook exists: {WORKBOOK.exists()}")
if not WORKBOOK.exists():
    raise SystemExit(2)

wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
try:
    ws = wb.active
    headers = [c.value for c in ws[2]]
    print(f"Sheet     : {ws.title}")
    print(f"Rows      : {ws.max_row}")
    print(f"Columns   : {ws.max_column}")
    print(f"Drone ID  : column B = {headers[1] if len(headers) > 1 else None!r}")
    kuin_column = headers.index("KUIN-G") + 1 if "KUIN-G" in headers else None
    print(f"KUIN-G    : column {kuin_column} = {'KUIN-G' if kuin_column else None!r}")
    if len(headers) < 2 or headers[1] != "Drone ID" or kuin_column is None:
        raise SystemExit("ERROR: Sample.xlsx schema is not compatible with the application.")
finally:
    wb.close()

QR_DIR.mkdir(parents=True, exist_ok=True)
print(f"QR dir    : {QR_DIR}")
print("Health check passed.")

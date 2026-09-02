#!/usr/bin/env python3
"""Serve the dashboard and persist asset entries to Excel + QR files."""

from __future__ import annotations

import base64
import datetime
import json
import os
import re
import shutil
import sys
import threading
import tempfile
from pathlib import Path
from urllib.parse import unquote, urlsplit
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer

from openpyxl import Workbook, load_workbook

APP_ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parents[1]
RESOURCE_ROOT = Path(getattr(sys, "_MEIPASS", APP_ROOT))
# Keep the live database beside the application by default.  This matches the
# deployment README and prevents the packaged build from silently writing to
# %LOCALAPPDATA% while the operator checks input/Sample.xlsx beside the EXE.
configured_data_root = os.environ.get("KUIN_G_DATA_ROOT")
DATA_ROOT = Path(configured_data_root) if configured_data_root else APP_ROOT
os.environ.setdefault("KUIN_G_DATA_ROOT", str(DATA_ROOT))
ROOT = DATA_ROOT
sys.path.insert(0, str(APP_ROOT))

from src.generate_UIDS import (
    create_qr_png,
    create_qr_svg,
    make_kuin,
    load_secret_key,
    remove_orphan_qr_files,
    write_distributable_workbook,
    write_qr_register,
    write_qr_stl_backup,
)
from tools.create_dashboard_data import load_records, write_outputs
from tools.data_protection import protect_data_before_startup, read_app_version

WORKBOOK = DATA_ROOT / "input" / "Sample.xlsx"
MASTER = ROOT / "output" / "master_register.xlsx"
QR_DIR = ROOT / "qr_codes"
REGISTER = ROOT / "output" / "qr_register.csv"
UPLOAD_DIR = DATA_ROOT / "dashboard" / "uploads"
PORT = 8765
PASSWORD_FILE = DATA_ROOT / "passwords.json"
SECRET_KEY_FILE = DATA_ROOT / "secret.key"
ARCHIVE_DIR = DATA_ROOT / "archive" / "deleted_records"
APP_VERSION = read_app_version(RESOURCE_ROOT)

# ThreadingHTTPServer can receive two saves at once. Excel workbooks are not a
# database and concurrent open/save operations can overwrite each other or
# leave a partially-written file. Serialize all workbook mutations.
WORKBOOK_LOCK = threading.RLock()


def prepare_runtime_data() -> None:
    APP_ROOT.mkdir(parents=True, exist_ok=True)
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    if not os.access(DATA_ROOT, os.W_OK):
        raise PermissionError(f"KUIN-G data folder is not writable: {DATA_ROOT}")

    backup_path = protect_data_before_startup(DATA_ROOT, RESOURCE_ROOT)
    if backup_path:
        print(f"Data snapshot created before {backup_path.stem.split('-')[-2]}: {backup_path}")

    # First-run initialization: copy bundled database/templates only when the
    # live database does not already exist.
    for relative_path in (
        Path("input") / "Sample.xlsx",
        Path("output") / "master_register.xlsx",
        Path("output") / "distributable_register.xlsx",
        Path("output") / "qr_register.csv",
    ):
        destination = DATA_ROOT / relative_path
        candidates = [
            APP_ROOT / relative_path,
            RESOURCE_ROOT / relative_path,
        ]
        if not destination.exists():
            for bundled in candidates:
                if bundled.exists() and bundled.resolve() != destination.resolve():
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(bundled, destination)
                    break

    for directory in (
        DATA_ROOT / "input",
        DATA_ROOT / "output",
        DATA_ROOT / "qr_codes",
        DATA_ROOT / "qr_stl_backup",
        DATA_ROOT / "dashboard" / "uploads",
        ARCHIVE_DIR,
    ):
        directory.mkdir(parents=True, exist_ok=True)

    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Database workbook not found: {WORKBOOK}")

    # Create/load the installation's 256-bit HMAC key once. Existing key files
    # are never overwritten, so KUIN-G values remain deterministic for this installation.
    load_secret_key(SECRET_KEY_FILE)


def archive_deleted_record(record: dict, qr_dir: Path, stl_dir: Path) -> Path:
    stamp = datetime.datetime.now(datetime.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    folder = ARCHIVE_DIR / f"{stamp}_{record['KUIN-G']}"
    folder.mkdir(parents=True, exist_ok=False)
    (folder / "record.json").write_text(json.dumps(record, indent=2, default=str), encoding="utf-8")
    for directory, suffix in ((qr_dir, ".png"), (qr_dir, ".svg"), (stl_dir, ".stl")):
        source = directory / f"{record['KUIN-G']}{suffix}"
        if source.is_file():
            shutil.copy2(source, folder / source.name)
    return folder


def dashboard_root() -> Path:
    for candidate in (APP_ROOT / "dashboard", RESOURCE_ROOT / "dashboard"):
        if (candidate / "index.html").is_file():
            return candidate
    raise FileNotFoundError(
        "Dashboard files are missing. Extract the complete KUIN-G folder before running KUIN-G.exe."
    )


def refresh_dashboard_data():
    # load_records() reads the same workbook that the UI writes.
    write_outputs(load_records())


def atomic_workbook_save(workbook, destination: Path) -> None:
    """Save an xlsx atomically, avoiding half-written database files."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{destination.stem}-", suffix=".xlsx", dir=str(destination.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, destination)
    except PermissionError as error:
        raise PermissionError(
            f"Cannot update {destination.name}. Close the workbook in Excel/LibreOffice and try again."
        ) from error
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except OSError:
            pass


def verify_saved_record(drone_id: str, kuin: str) -> None:
    """Reopen the workbook and verify the exact row was committed."""
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in range(3, ws.max_row + 1):
            if str(ws.cell(row, 2).value or "") == drone_id:
                saved_kuin = str(ws.cell(row, 28).value or "")
                if saved_kuin != kuin:
                    raise IOError(
                        f"Excel verification failed for {drone_id}: expected KUIN-G {kuin}, found {saved_kuin or '<blank>'}."
                    )
                return
    finally:
        wb.close()
    raise IOError(f"Excel verification failed: {drone_id} was not found after save.")


def rebuild_registers_from_workbook() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[2]]
        records = [
            dict(zip(headers, values))
            for values in worksheet.iter_rows(min_row=3, values_only=True)
            if len(values) > 1 and values[1] not in (None, "")
        ]
    finally:
        workbook.close()

    kuin_values = [record["KUIN-G"] for record in records]
    master = Workbook()
    master_sheet = master.active
    master_sheet.title = "Master Register"
    master_sheet.append(["Drone ID", "KUIN-G"])
    qr_records = []
    for record in records:
        kuin = record["KUIN-G"]
        master_sheet.append([record["Drone ID"], kuin])
        qr_records.append({
            "Drone ID": record["Drone ID"],
            "KUIN-G": kuin,
            "QR PNG File": f"{kuin}.png",
            "QR SVG File": f"{kuin}.svg",
            "QR STL File": f"{kuin}.stl",
        })
    MASTER.parent.mkdir(parents=True, exist_ok=True)
    master.save(MASTER)
    master.close()

    QR_DIR.mkdir(parents=True, exist_ok=True)
    write_distributable_workbook(ROOT / "output" / "distributable_register.xlsx", kuin_values, QR_DIR)
    remove_orphan_qr_files(QR_DIR, kuin_values)
    write_qr_stl_backup(ROOT / "qr_stl_backup", kuin_values)
    write_qr_register(REGISTER, qr_records)
    refresh_dashboard_data()


def load_passwords():
    if PASSWORD_FILE.exists():
        return json.loads(PASSWORD_FILE.read_text(encoding="utf-8"))
    passwords = {
        "admin": os.environ.get("KUIN_ADMIN_PASSWORD", "admin"),
        "user": os.environ.get("KUIN_USER_PASSWORD", "user"),
    }
    PASSWORD_FILE.write_text(json.dumps(passwords, indent=2), encoding="utf-8")
    return passwords


def save_passwords(passwords):
    PASSWORD_FILE.write_text(json.dumps(passwords, indent=2), encoding="utf-8")


def fragment(value, length):
    letters = re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()
    return (letters or "UNK")[:length].ljust(length, "X")


def hierarchy_id(record, existing_ids):
    base = "-".join([
        fragment(record["Command"], 2), fragment(record["Corps"], 2),
        fragment(record["Division"], 3), fragment(record["Brigade"], 4), fragment(record["Unit"], 4),
        fragment(record["Drone Name"], 3), fragment(record["Type"], 3),
        f"{int(record.get('Ser No') or 1):02d}",
    ])
    matches = sum(existing_id == base or existing_id.startswith(f"{base}-") for existing_id in existing_ids)
    return base if matches == 0 else f"{base}-{matches + 1:02d}"


class DashboardHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(dashboard_root()), **kwargs)

    def log_message(self, format, *args):
        print(f"[{self.log_date_time_string()}] {format % args}")

    def end_headers(self):
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        super().end_headers()

    def do_GET(self):
        request_path = urlsplit(self.path).path
        if request_path.startswith("/qr/"):
            filename = unquote(request_path.removeprefix("/qr/")).replace("\\", "/")
            if "/" in filename or not re.fullmatch(r"[A-Za-z0-9_-]+\.(?:png|svg)", filename):
                self.send_error(404)
                return
            qr_path = (QR_DIR / filename).resolve()
            if not qr_path.is_file() or QR_DIR.resolve() not in qr_path.parents:
                self.send_error(404)
                return
            body = qr_path.read_bytes()
            content_type = "image/svg+xml" if qr_path.suffix.lower() == ".svg" else "image/png"
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(body)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(body)
            return
        if request_path == "/data.js":
            body = (
                "window.dashboardData = "
                + json.dumps(load_records(), ensure_ascii=True, separators=(",", ":"))
                + ";\n"
            ).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/javascript; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        super().do_GET()

    def do_POST(self):
        if self.path == "/api/login":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                request = json.loads(self.rfile.read(length))
                passwords = load_passwords()
                valid = request.get("role") in passwords and request.get("password") == passwords[request["role"]]
                self._send_json(200 if valid else 401, {"authenticated": valid})
            except Exception as error:
                self._send_json(400, {"error": str(error)})
            return

        if self.path == "/api/password":
            try:
                role = self.headers.get("X-Role", "")
                passwords = load_passwords()
                if role not in passwords or self.headers.get("X-Password") != passwords[role]:
                    raise PermissionError("Current password is incorrect")
                length = int(self.headers.get("Content-Length", "0"))
                request = json.loads(self.rfile.read(length))
                new_password = str(request.get("new_password", ""))
                if len(new_password) < 4:
                    raise ValueError("New password must contain at least 4 characters")
                passwords[role] = new_password
                save_passwords(passwords)
                self._send_json(200, {"changed": True})
            except Exception as error:
                self._send_json(403 if isinstance(error, PermissionError) else 400, {"error": str(error)})
            return

        if self.path == "/api/assets/edit":
            try:
                passwords = load_passwords()
                if self.headers.get("X-Role") != "admin" or self.headers.get("X-Password") != passwords["admin"]:
                    raise PermissionError("Admin role required")
                length = int(self.headers.get("Content-Length", "0"))
                request = json.loads(self.rfile.read(length))
                drone_id = request.get("Drone ID", "")
                changes = request.get("changes", {})
                forbidden = {field for field in changes if field in {"Drone ID", "Ser No", "Command", "Corps", "Brigade", "Unit", "Drone Name", "Type"}}
                if forbidden:
                    raise ValueError("Primary-key fields cannot be edited")
                with WORKBOOK_LOCK:
                    workbook = load_workbook(WORKBOOK)
                    try:
                        worksheet = workbook.active
                        row_number = next((row for row in range(3, worksheet.max_row + 1) if str(worksheet.cell(row, 2).value) == drone_id), None)
                        if row_number is None:
                            raise ValueError("Drone ID not found")
                        headers = [cell.value for cell in worksheet[2]]
                        for field, value in changes.items():
                            if field not in headers:
                                raise ValueError(f"Unknown field: {field}")
                            worksheet.cell(row_number, headers.index(field) + 1, value)
                        atomic_workbook_save(workbook, WORKBOOK)
                    finally:
                        workbook.close()
                    rebuild_registers_from_workbook()
                self._send_json(200, {"updated": drone_id, "changes": changes})
            except Exception as error:
                self._send_json(403 if isinstance(error, PermissionError) else 400, {"error": str(error)})
            return

        if self.path != "/api/assets":
            self.send_error(404)
            return

        try:
            role = self.headers.get("X-Role", "")
            passwords = load_passwords()
            if role not in passwords or self.headers.get("X-Password") != passwords[role]:
                raise PermissionError("Valid User or Admin login required")

            length = int(self.headers.get("Content-Length", "0"))
            record = json.loads(self.rfile.read(length))
            record.pop("Drone ID", None)
            record.pop("KUIN-G", None)

            with WORKBOOK_LOCK:
                workbook = load_workbook(WORKBOOK)
                try:
                    worksheet = workbook.active
                    headers = [cell.value for cell in worksheet[2]]
                    if len(headers) < 28 or headers[1] != "Drone ID" or headers[27] != "KUIN-G":
                        raise ValueError(
                            "Sample.xlsx has an unexpected schema. Expected Drone ID in column B and KUIN-G in column AB."
                        )
                    required = [
                        header for header in headers
                        if header not in ("Ser No", "Drone ID", "KUIN-G", "Payload Description", "Image Front", "Image Back", "Image Top", "Image Bottom")
                    ]
                    missing = [header for header in required if str(record.get(header, "")).strip() == ""]
                    if missing:
                        raise ValueError(f"Missing fields: {', '.join(missing)}")

                    existing_ids = {
                        str(worksheet.cell(row=row, column=2).value).strip()
                        for row in range(3, worksheet.max_row + 1)
                        if worksheet.cell(row=row, column=2).value not in (None, "")
                    }
                    record["Ser No"] = max(
                        [int(worksheet.cell(row=row, column=1).value or 0) for row in range(3, worksheet.max_row + 1)] or [0]
                    ) + 1
                    record["Drone ID"] = hierarchy_id(record, existing_ids)
                    # KUIN-G = first 16 Base32 characters of HMAC-SHA256(secret_key, Drone ID).
                    # The secret is 256 bits and is loaded from the protected data-root key file
                    # (or from KUIN_G_SECRET_KEY_B64 / KUIN_G_SECRET_KEY_FILE).
                    secret_key = load_secret_key(SECRET_KEY_FILE)
                    record["KUIN-G"] = make_kuin(record["Drone ID"], secret_key)

                    # Save uploaded images before the Excel transaction. If the
                    # workbook fails, they remain harmless files and can be
                    # cleaned up manually; the database row is never falsely reported as saved.
                    for image_field in ("Image Front", "Image Back", "Image Top", "Image Bottom"):
                        image_data = record.get(image_field, "")
                        if not image_data:
                            continue
                        match = re.fullmatch(r"data:(image/png|image/jpeg);base64,(.+)", image_data)
                        if not match:
                            raise ValueError("Image must be PNG or JPEG")
                        raw_image = base64.b64decode(match.group(2), validate=True)
                        if len(raw_image) > 2 * 1024 * 1024:
                            raise ValueError("Image must be 2 MB or smaller")
                        extension = "png" if match.group(1) == "image/png" else "jpg"
                        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
                        image_name = f"{record['Drone ID']}-{image_field.split()[-1].lower()}.{extension}"
                        (UPLOAD_DIR / image_name).write_bytes(raw_image)
                        record[image_field] = f"uploads/{image_name}"

                    row_number = worksheet.max_row + 1
                    for column, header in enumerate(headers, start=1):
                        worksheet.cell(row=row_number, column=column, value=record.get(header, ""))

                    # The actual database commit happens here.
                    atomic_workbook_save(workbook, WORKBOOK)
                finally:
                    workbook.close()

                # Never claim success until the saved workbook is reopened and
                # the new row is confirmed.
                verify_saved_record(record["Drone ID"], record["KUIN-G"])

                # Generate the QR only after Excel has committed successfully.
                QR_DIR.mkdir(parents=True, exist_ok=True)
                png_path = QR_DIR / f"{record['KUIN-G']}.png"
                svg_path = QR_DIR / f"{record['KUIN-G']}.svg"
                create_qr_png(record["KUIN-G"], png_path)
                create_qr_svg(record["KUIN-G"], svg_path)

                # Rebuild the derived registers and dashboard from the newly
                # committed workbook, so every view has the same source of truth.
                rebuild_registers_from_workbook()

            response_record = dict(record)
            response_record["Serv"] = "Svc" if record["Serv"] == "Ser" else "Unsvc"
            response_record["QR PNG"] = f"/qr/{record['KUIN-G']}.png"
            response_record["QR SVG"] = f"/qr/{record['KUIN-G']}.svg"
            self._send_json(
                201,
                {
                    "record": response_record,
                    "kuin": record["KUIN-G"],
                    "qr_png": f"/qr/{record['KUIN-G']}.png",
                    "qr_svg": f"/qr/{record['KUIN-G']}.svg",
                    "excel": str(WORKBOOK),
                },
            )
        except Exception as error:
            print(f"SAVE ERROR: {type(error).__name__}: {error}")
            self._send_json(403 if isinstance(error, PermissionError) else 400, {"error": str(error)})

    def do_DELETE(self):
        if not self.path.startswith("/api/assets/"):
            self.send_error(404)
            return
        try:
            if self.headers.get("X-Role", "user") != "admin" or self.headers.get("X-Password") != load_passwords()["admin"]:
                raise PermissionError("Admin role required")
            drone_id = unquote(self.path.removeprefix("/api/assets/"))
            with WORKBOOK_LOCK:
                workbook = load_workbook(WORKBOOK)
                try:
                    worksheet = workbook.active
                    row_to_delete = next((row for row in range(3, worksheet.max_row + 1) if str(worksheet.cell(row, 2).value) == drone_id), None)
                    if row_to_delete is None:
                        raise ValueError("Drone ID not found")
                    headers = [cell.value for cell in worksheet[2]]
                    deleted_record = dict(zip(headers, next(worksheet.iter_rows(min_row=row_to_delete, max_row=row_to_delete, values_only=True))))
                    archive_deleted_record(deleted_record, QR_DIR, DATA_ROOT / "qr_stl_backup")
                    worksheet.delete_rows(row_to_delete)
                    atomic_workbook_save(workbook, WORKBOOK)
                finally:
                    workbook.close()
                rebuild_registers_from_workbook()
            self._send_json(200, {"deleted": drone_id})
        except Exception as error:
            self._send_json(403 if isinstance(error, PermissionError) else 400, {"error": str(error)})

    def _send_json(self, status, payload):
        body = json.dumps(payload, ensure_ascii=True).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)


if __name__ == "__main__":
    prepare_runtime_data()
    refresh_dashboard_data()
    print(f"KUIN-G dashboard: http://127.0.0.1:{PORT}/index.html")
    print(f"KUIN-G Excel database: {WORKBOOK}")
    print(f"KUIN-G QR directory: {QR_DIR}")
    ThreadingHTTPServer(("127.0.0.1", PORT), DashboardHandler).serve_forever()

#!/usr/bin/env python3
"""Create the dashboard dataset and a Power BI-friendly CSV export."""

from __future__ import annotations

import csv
import json
import os
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


PROJECT_DIR = Path(os.environ.get("KUIN_G_DATA_ROOT", Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parents[1]))
INPUT_FILE = PROJECT_DIR / "input" / "Sample.xlsx"
DATA_FILE = PROJECT_DIR / "dashboard" / "data.js"
CSV_FILE = PROJECT_DIR / "dashboard" / "fleet_register.csv"

EXPORT_FIELDS = [
    "Ser No", "Drone ID", "Drone Name", "Type", "Form Factor", "OEM",
    "Range", "Weight (KG)", "Endurance (min)", "Day/Night Capability", "Payload", "Payload Weight",
    "Payload Description", "Guidance", "Anti Ew", "C2 Link Frequency", "Proc Fund",
    "Serv", "Cost (in Thousands)", "Image Front", "Image Back", "Image Top", "Image Bottom",
    "Unit", "Brigade", "Division", "Corps", "Command", "KUIN-G",
]


def load_records() -> list[dict[str, object]]:
    workbook = load_workbook(INPUT_FILE, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[2]]
        return [
            dict(zip(headers, values))
            for values in worksheet.iter_rows(min_row=3, values_only=True)
            if values[1] not in (None, "")
        ]
    finally:
        workbook.close()


def write_outputs(records: list[dict[str, object]]) -> None:
    DATA_FILE.write_text(
        "window.dashboardData = "
        + json.dumps(records, ensure_ascii=True, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )

    with CSV_FILE.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=EXPORT_FIELDS)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    records = load_records()
    write_outputs(records)
    print(f"Dashboard data: {DATA_FILE}")
    print(f"Power BI CSV:    {CSV_FILE}")
    print(f"Records:         {len(records)}")
    print(f"Serviceability:  {dict(Counter(row['Serv'] for row in records))}")


if __name__ == "__main__":
    main()
"""Create an empty KUIN-G input workbook with dropdown validation."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path("input") / "Sample.xlsx"
HEADERS = [
    "Ser No", "Drone ID", "Drone Name", "Type", "Form Factor", "OEM", "Range",
    "Weight (KG)", "Endurance (min)", "Day/Night Capability", "Payload", "Payload Weight", "Payload Description",
    "Guidance", "Anti Ew", "C2 Link Frequency", "Proc Fund", "Serv",
    "Cost (in Thousands)", "Image Front", "Image Back", "Image Top", "Image Bottom",
    "Unit", "Brigade", "Division", "Corps", "Command", "KUIN-G",
]
CONTROLLED_VALUES = {
    "Type": ["Trg", "Svl (SR)", "Svl (MR)", "FPV", "Kamikaze", "Lgs", "Loitering Munition"],
    "Form Factor": ["QC", "HC", "Fixed Wg", "FIxed Wing VTOL", "Swarm"],
    "Range": ["< 5 km", "5-10 km", "10-30 km", "31-100 km", "> 100 km"],
    "Day/Night Capability": ["Day", "Night"],
    "C2 Link Frequency": ["1.4 GHz", "2.4 GHz", "5.8 GHz", "900 MHz"],
    "Proc Fund": ["ATG", "Regtl", "Unit", "Central"],
    "Anti Ew": ["Nil"],
    "Serv": ["Ser", "Unser"],
}


def add_dropdown(worksheet, values, column):
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
    worksheet.add_data_validation(validation)
    validation.add(f"{column}3:{column}1002")


def main():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Drone Inventory - KUIN-G Input"
    worksheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    for column, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=2, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        worksheet.column_dimensions[cell.column_letter].width = 20
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:{worksheet.cell(row=2, column=len(HEADERS)).column_letter}1002"
    for field, values in CONTROLLED_VALUES.items():
        column = worksheet.cell(row=2, column=HEADERS.index(field) + 1).column_letter
        add_dropdown(worksheet, values, column)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT)
    print(f"Wrote empty KUIN-G workbook: {OUTPUT}")


if __name__ == "__main__":
    main()

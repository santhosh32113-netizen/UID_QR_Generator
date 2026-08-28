#!/usr/bin/env python3
"""Migrate the sample workbook to the hierarchy-based asset schema."""

from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
import sys
sys.path.insert(0, str(ROOT))
from src.generate_UIDS import make_kuin, remove_orphan_qr_files, write_distributable_workbook, write_qr_register

SOURCE = ROOT / "input" / "Sample.xlsx"
HEADERS = [
    "Ser No", "Drone ID", "Drone Name", "Type", "Form Factor", "OEM", "Range",
    "Weight (KG)", "Endurance (min)", "Payload", "Payload Weight", "Payload Description",
    "Guidance", "Anti Ew", "C2 Link Frequency", "Proc Fund", "Serv",
    "Cost (in Thousands)", "Image Front", "Image Back", "Image Top", "Image Bottom",
    "Unit", "Brigade", "Division", "Corps", "Command", "KUIN-G",
]
CONTROLLED_VALUES = {
    "Type": ["Trg", "Svl (SR)", "Svl (MR)", "FPV", "Kamikaze", "Lgs", "Loitering Munition"],
    "Form Factor": ["QC", "HC", "Fixed Wg", "FIxed Wing VTOL", "Swarm"],
    "Range": ["< 5 km", "5-10 km", "10-30 km", "31-100 km", "> 100 km"],
    "C2 Link Frequency": ["1.4 GHz", "2.4 GHz", "5.8 GHz", "900 MHz"],
    "Proc Fund": ["ATG", "Regtl", "Unit", "Central"],
    "Serv": ["Ser", "Unser"],
}


def fragment(value: object, length: int) -> str:
    letters = re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()
    return (letters or "UNK")[:length].ljust(length, "X")


def hierarchy_id(row: dict[str, object], occurrence: int) -> str:
    base = "-".join([
        fragment(row["Command"], 2), fragment(row["Corps"], 2),
        fragment(row["Brigade"], 4), fragment(row["Unit"], 4),
        fragment(row["Drone Name"], 3), fragment(row["Type"], 3),
        f"{int(row['Ser No'] or 0):02d}",
    ])
    return base if occurrence == 1 else f"{base}-{occurrence:02d}"


def normalize_range(value: object) -> str:
    if isinstance(value, (int, float)):
        if value < 5:
            return "< 5 km"
        if value <= 10:
            return "5-10 km"
        if value <= 30:
            return "10-30 km"
        if value <= 100:
            return "31-100 km"
        return "> 100 km"
    return str(value or "< 5 km")


def normalize_endurance(value: object) -> int:
    match = re.search(r"[0-9]+(?:\.[0-9]+)?", str(value or "0"))
    number = float(match.group()) if match else 0
    return int(number * 60 if "hr" in str(value).lower() else number)


def main() -> None:
    source_wb = load_workbook(SOURCE, read_only=True, data_only=True)
    source_ws = source_wb.active
    old_headers = [cell.value for cell in source_ws[2]]
    old_rows = [dict(zip(old_headers, values)) for values in source_ws.iter_rows(min_row=3, values_only=True) if values[1]]
    source_wb.close()

    occurrences: defaultdict[str, int] = defaultdict(int)
    rows = []
    for old in old_rows:
        row = {header: old.get(header, "") for header in old_headers}
        row["Range"] = normalize_range(old.get("Range") or old.get("Range (KM)"))
        row["Endurance (min)"] = normalize_endurance(old.get("Endurance (min)") or old.get("Endurance"))
        row["C2 Link Frequency"] = old.get("C2 Link Frequency") or old.get("Link Freq") or ""
        row["Payload"] = old.get("Payload", "")
        row["Payload Weight"] = old.get("Payload Weight", "")
        row["Guidance"] = old.get("Guidance", "")
        row["Anti Ew"] = old.get("Anti Ew", "")
        row["Unit"] = old.get("Unit", "")
        legacy_brigade = old.get("Brigade") or old.get("Fmn") or "Brigade 1"
        row["Brigade"] = re.sub(r"^Fmn\b", "Brigade", str(legacy_brigade), flags=re.IGNORECASE)
        row["Division"] = old.get("Division") or "Division Alpha"
        row["Corps"] = old.get("Corps") or "Corps North"
        row["Command"] = old.get("Command") or "Command East"
        row["Serv"] = "Ser" if str(old.get("Serv", "")).strip().lower() in {"svc", "ser", "serviceable"} else "Unser"
        old_image = old.get("Image", "")
        row["Image Front"] = old.get("Image Front") or old_image
        row["Image Back"] = old.get("Image Back", "")
        row["Image Top"] = old.get("Image Top", "")
        row["Image Bottom"] = old.get("Image Bottom", "")
        base = "-".join([fragment(row["Command"], 2), fragment(row["Corps"], 2), fragment(row["Brigade"], 4), fragment(row["Unit"], 4), fragment(row["Drone Name"], 3), fragment(row["Type"], 3), f"{int(row['Ser No'] or 0):02d}"])
        occurrences[base] += 1
        row["Drone ID"] = hierarchy_id(row, occurrences[base])
        row["KUIN-G"] = make_kuin(row["Drone ID"])
        rows.append([row.get(header, "") for header in HEADERS])

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "Drone Inventory - Schema Migrated"
    for column, header in enumerate(HEADERS, start=1):
        worksheet.cell(row=2, column=column, value=header)
    for row_number, values in enumerate(rows, start=3):
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=row_number, column=column, value=value)
    worksheet.freeze_panes = "A3"
    worksheet.auto_filter.ref = f"A2:X{worksheet.max_row}"
    for column in range(1, len(HEADERS) + 1):
        worksheet.column_dimensions[worksheet.cell(row=2, column=column).column_letter].width = 20
    validation_values = dict(CONTROLLED_VALUES)
    for field in ("OEM", "Unit", "Brigade", "Division", "Corps", "Command"):
        validation_values[field] = sorted({str(row[HEADERS.index(field)]) for row in rows if row[HEADERS.index(field)]})
    for field, values in validation_values.items():
        column = HEADERS.index(field) + 1
        validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=False)
        worksheet.add_data_validation(validation)
        validation.add(f"{worksheet.cell(row=3, column=column).column_letter}3:{worksheet.cell(row=3, column=column).column_letter}{worksheet.max_row}")
    workbook.save(SOURCE)
    qr_dir = ROOT / "qr_codes"
    kuin_values = [row[-1] for row in rows]
    master = Workbook()
    master_sheet = master.active
    master_sheet.title = "Master Register"
    master_sheet.append(["Drone ID", "KUIN-G"])
    qr_records = []
    for values in rows:
        kuin = values[-1]
        master_sheet.append([values[1], kuin])
        qr_records.append({"Drone ID": values[1], "KUIN-G": kuin,
                   "QR PNG File": f"{kuin}.png", "QR SVG File": f"{kuin}.svg"})
    master.save(ROOT / "output" / "master_register.xlsx")
    write_distributable_workbook(ROOT / "output" / "distributable_register.xlsx", kuin_values, qr_dir)
    remove_orphan_qr_files(qr_dir, kuin_values)
    write_qr_register(ROOT / "output" / "qr_register.csv", qr_records)
    print(f"Migrated {len(rows)} records to {len(HEADERS)} fields")


if __name__ == "__main__":
    main()